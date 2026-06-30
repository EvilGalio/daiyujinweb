"""Freight importer v2: parses all rate types from D重量运费.xlsx.

Rate types:
  - small_matrix: zone-expanded matrix from 区域运费DHL / 区域运费FedEX
  - document: fixed prices from 2KG内文件
  - heavy_per_kg: per-kg unit prices from 广诚DHL运费 / 广诚FedEx运费
  - zone_mapping: country → zone/code mappings
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = PROJECT_ROOT / "_private" / "source_data" / "freight" / "D重量运费.xlsx"
if not WORKBOOK.exists():
    WORKBOOK = PROJECT_ROOT / "D重量运费.xlsx"


def _num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _text(v: Any) -> str:
    return str(v).strip() if v is not None else ""


# ═══════════════════════════════════════════════════
# Zone mappings: country → zone_code per carrier
# ═══════════════════════════════════════════════════

def parse_zone_mappings(ws, carrier: str) -> list[dict[str, Any]]:
    """Extract country → zone_code from a matrix sheet."""
    import re
    zones = []
    for row_idx in range(2, ws.max_row + 1):
        country = _text(ws.cell(row=row_idx, column=1).value)
        if not country or "kg" in country.lower():
            continue
        # Skip weight tier rows that got mixed in (e.g. "21-44", "100-299")
        if re.match(r'^[\d.]+\s*-\s*[\d.]+$', country):
            continue
        zone_code = _text(ws.cell(row=row_idx, column=2).value)
        if not zone_code:
            continue
        zones.append({
            "carrier": carrier,
            "country_cn": country,
            "zone_code": zone_code,
            "source_sheet": ws.title,
            "source_row": row_idx,
        })
    return zones


# ═══════════════════════════════════════════════════
# Small matrix: zone-expanded matrix (existing parser enhanced)
# ═══════════════════════════════════════════════════

def parse_small_matrix(ws, carrier: str) -> list[dict[str, Any]]:
    """Parse zone-expanded matrix like 区域运费DHL / 区域运费FedEX."""
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header:
        return []
    row_vals = header[0]
    weight_columns: list[tuple[int, float]] = []
    for col_idx, heading in enumerate(row_vals):
        w = _num(heading)
        if w is not None and w > 0:
            weight_columns.append((col_idx, w))

    records = []
    for row_idx in range(2, ws.max_row + 1):
        country = _text(ws.cell(row=row_idx, column=1).value)
        if not country or "kg" in country.lower():
            continue
        zone_code = _text(ws.cell(row=row_idx, column=2).value)
        if not zone_code:
            continue

        for col_idx, charge_weight in weight_columns:
            price = _num(ws.cell(row=row_idx, column=col_idx + 1).value)
            if price is None:
                continue
            records.append({
                "carrier": carrier,
                "country_cn": country,
                "zone_code": zone_code,
                "cargo_type": "package",
                "pricing_mode": "small_matrix",
                "weight_min": charge_weight,
                "weight_max": charge_weight,
                "charge_weight": charge_weight,
                "currency": "CNY",
                "price_type": "fixed",
                "price": price,
                "source_sheet": ws.title,
                "source_row": row_idx,
            })
    return records


# ═══════════════════════════════════════════════════
# Document rates: 2KG内文件
# ═══════════════════════════════════════════════════

def parse_document_rates(wb) -> list[dict[str, Any]]:
    """Parse DHL document rates from 2KG内文件 sheet."""
    ws = wb["2KG内文件"]
    records = []
    # Row 3: header (KG, 1区, 2区, ... 9区)
    # Rows 4-7: data
    weight_tiers = [0.5, 1.0, 1.5, 2.0]
    for data_row, wt in enumerate(weight_tiers, start=4):
        for zone_num in range(1, 10):
            price = _num(ws.cell(row=data_row, column=zone_num + 1).value)
            if price is None:
                continue
            records.append({
                "carrier": "DHL",
                "cargo_type": "document",
                "pricing_mode": "document",
                "zone_code": str(zone_num),
                "weight_min": wt,
                "weight_max": wt,
                "charge_weight": wt,
                "currency": "CNY",
                "price_type": "fixed",
                "price": price,
                "source_sheet": "2KG内文件",
                "source_row": data_row,
            })
    return records


# ═══════════════════════════════════════════════════
# Heavy cargo: 广诚DHL运费
# ═══════════════════════════════════════════════════

def parse_dhl_heavy_rates(ws) -> list[dict[str, Any]]:
    """Parse heavy per-kg rates from 广诚DHL运费 (rows 64+)."""
    records = []
    # Heavy section starts after "30.1 KG起" marker
    # Row 65: header (自, 至, 1区, 2区, ... 9区)
    # Row 66+: data rows
    for row_idx in range(66, ws.max_row + 1):
        lo = _num(ws.cell(row=row_idx, column=1).value)
        hi = _num(ws.cell(row=row_idx, column=2).value)
        if lo is None:
            continue
        for zone_num in range(1, 10):
            price = _num(ws.cell(row=row_idx, column=2 + zone_num).value)
            if price is None:
                continue
            records.append({
                "carrier": "DHL",
                "cargo_type": "package",
                "pricing_mode": "heavy_per_kg",
                "zone_code": str(zone_num),
                "weight_min": lo,
                "weight_max": hi,
                "charge_weight": None,
                "currency": "CNY",
                "price_type": "per_kg",
                "price": price,
                "source_sheet": "广诚DHL运费",
                "source_row": row_idx,
            })
    return records


# ═══════════════════════════════════════════════════
# Heavy cargo: 广诚FedEx运费
# ═══════════════════════════════════════════════════

def parse_fedex_heavy_rates(ws) -> list[dict[str, Any]]:
    """Parse heavy per-kg rates from 广诚FedEx运费 (rows 44+)."""
    # Row 2: country names
    # Row 3: zone codes (letters)
    # Row 46+: heavy tier rows
    zone_headers = {}
    for col_idx in range(2, ws.max_column + 1):
        zone = _text(ws.cell(row=3, column=col_idx).value)
        if zone and zone.isalnum():
            zone_headers[col_idx] = zone

    records = []
    for row_idx in range(46, ws.max_row + 1):
        lo_text = _text(ws.cell(row=row_idx, column=1).value)
        if not lo_text or "重货" in lo_text:
            continue
        # Parse "21.0 - 44.0" style
        parts = lo_text.replace("-", " ").replace("–", " ").split()
        try:
            lo = float(parts[0])
            hi = float(parts[1]) if len(parts) > 1 else None
        except (ValueError, IndexError):
            continue

        for col_idx, zone_code in zone_headers.items():
            price = _num(ws.cell(row=row_idx, column=col_idx).value)
            if price is None:
                continue
            records.append({
                "carrier": "FedEx",
                "cargo_type": "package",
                "pricing_mode": "heavy_per_kg",
                "zone_code": zone_code,
                "weight_min": lo,
                "weight_max": hi,
                "charge_weight": None,
                "currency": "CNY",
                "price_type": "per_kg",
                "price": price,
                "source_sheet": "广诚FedEx运费",
                "source_row": row_idx,
            })
    return records


# ═══════════════════════════════════════════════════
# Main import orchestrator
# ═══════════════════════════════════════════════════

def import_v2() -> dict[str, Any]:
    wb = load_workbook(WORKBOOK, read_only=False, data_only=True)
    try:
        zones: list[dict[str, Any]] = []
        rate_cards: list[dict[str, Any]] = []

        # ── Small matrix ──
        for sheet_name, carrier in [("区域运费DHL", "DHL"), ("区域运费FedEX", "FedEx")]:
            ws = wb[sheet_name]
            zones.extend(parse_zone_mappings(ws, carrier))
            rate_cards.extend(parse_small_matrix(ws, carrier))

        # ── Document ──
        doc_rates = parse_document_rates(wb)
        rate_cards.extend(doc_rates)

        # ── Heavy DHL ──
        ws_gc_dhl = wb["广诚DHL运费"]
        dhl_heavy = parse_dhl_heavy_rates(ws_gc_dhl)
        rate_cards.extend(dhl_heavy)

        # ── Heavy FedEx ──
        ws_gc_fex = wb["广诚FedEx运费"]
        fex_heavy = parse_fedex_heavy_rates(ws_gc_fex)
        rate_cards.extend(fex_heavy)

        # ── Build report ──
        report = {
            "zone_count": len(zones),
            "rate_card_count": len(rate_cards),
            "small_matrix_count": sum(1 for r in rate_cards if r["pricing_mode"] == "small_matrix"),
            "document_rate_count": len(doc_rates),
            "dhl_heavy_count": len(dhl_heavy),
            "fedex_heavy_count": len(fex_heavy),
        }
        return {
            "report": report,
            "zones": zones,
            "rate_cards": rate_cards,
        }
    finally:
        wb.close()


if __name__ == "__main__":
    import json
    result = import_v2()
    print(json.dumps(result["report"], indent=2))
