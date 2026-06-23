from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def find_freight_workbook(root: Path) -> Path:
    matches = sorted(root.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError("No .xlsx freight workbook found in project root")
    return matches[0]


def _as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def _parse_matrix_sheet(ws, carrier: str) -> list[dict[str, Any]]:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    weight_columns: list[tuple[int, float]] = []
    for index, heading in enumerate(header):
        weight = _as_float(heading)
        if weight is not None:
            weight_columns.append((index, weight))

    records: list[dict[str, Any]] = []
    seen_data = False
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        country = row[0] if len(row) > 0 else None
        zone = row[1] if len(row) > 1 else None
        if not country:
            if seen_data:
                break
            continue
        country_text = str(country).strip()
        if "kg" in country_text.lower() or country_text.upper() == "KG":
            break
        if not country or zone is None:
            continue

        for col_index, weight_kg in weight_columns:
            if col_index >= len(row):
                continue
            price = _as_float(row[col_index])
            if price is None:
                continue
            records.append(
                {
                    "carrier": carrier,
                    "country": country_text,
                    "zone": str(zone).strip(),
                    "weight_kg": weight_kg,
                    "price": price,
                    "currency": "CNY",
                    "source_sheet": ws.title,
                    "source_row": row_index,
                }
            )
            seen_data = True
    return records


def parse_freight_workbook(workbook_path: str | Path, include_records: bool = True) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    # The workbook is small enough to load normally. Using read_only=True was
    # unreliable with cached formula values across multiple sheets in this file.
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        # Match the two zone-matrix sheets by name so the parser stays correct
        # even when sheets are reordered. Other freight sheets use different
        # layouts and will be handled by later importer versions.
        sheet_specs: list[tuple[str, str]] = []
        for ws in wb.worksheets:
            title = ws.title
            normalized = title.lower()
            if title.startswith("区域运费") and "dhl" in normalized:
                sheet_specs.append((title, "DHL"))
            elif title.startswith("区域运费") and "fedex" in normalized:
                sheet_specs.append((title, "FedEx"))
        records: list[dict[str, Any]] = []
        for sheet_name, carrier in sheet_specs:
            ws = wb[sheet_name]
            records.extend(_parse_matrix_sheet(ws, carrier))
    finally:
        wb.close()

    countries = sorted({record["country"] for record in records})
    carriers = sorted({record["carrier"] for record in records})
    weights = sorted({record["weight_kg"] for record in records})

    result: dict[str, Any] = {
        "workbook": str(workbook_path),
        "record_count": len(records),
        "country_count": len(countries),
        "carriers": carriers,
        "min_weight_kg": weights[0] if weights else None,
        "max_weight_kg": weights[-1] if weights else None,
        "sample_records": records[:5],
    }
    if include_records:
        result["records"] = records
    return result
