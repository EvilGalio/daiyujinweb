"""DHL-only importer for A重量运费重制版.xlsx Sheet1."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = PROJECT_ROOT / "A重量运费重制版.xlsx"


def _num(v: Any) -> float | None:
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s: return None
        try: return float(s)
        except: return None
    return None


def _text(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def import_dhl() -> dict[str, Any]:
    wb = load_workbook(WORKBOOK, read_only=False, data_only=True)
    ws = wb["Sheet1"]
    try:
        # ── Zones: A2:B234 ──
        zones = []
        for r in range(2, 235):
            country = _text(ws.cell(row=r, column=1).value)
            zone = _text(ws.cell(row=r, column=2).value)
            if not country or not zone:
                continue
            zones.append({"country_cn": country, "zone_code": zone, "row": r})

        # ── Small rates: C2:BJ234 (0.5-30kg) ──
        weight_headers = []
        for c in range(3, 63):
            w = _num(ws.cell(row=1, column=c).value)
            if w is not None:
                weight_headers.append((c, w))

        small_rates = []
        for r in range(2, 235):
            country = _text(ws.cell(row=r, column=1).value)
            zone = _text(ws.cell(row=r, column=2).value)
            if not country:
                continue
            for col, charge_w in weight_headers:
                price = _num(ws.cell(row=r, column=col).value)
                if price is not None:
                    small_rates.append({
                        "country_cn": country,
                        "zone_code": zone,
                        "charge_weight": charge_w,
                        "base_price_cny": price,
                    })

        # ── Heavy rates: A235:J238 ──
        heavy_tiers = []
        for r in range(236, 239):
            tier = _text(ws.cell(row=r, column=1).value)
            prices = {}
            for zn in range(1, 10):
                p = _num(ws.cell(row=r, column=1 + zn).value)
                if p is not None:
                    prices[f"{zn}区"] = p
            if tier:
                heavy_tiers.append({"tier": tier, "zones": prices})

        # ── Config ──
        config = {
            "default_currency": "USD",
            "small_multiplier": 1.8,
            "heavy_multiplier": 1.7,
            "usd_divisor": 6,
            "eur_divisor": 7,
            "small_weight_limit_kg": 33,
        }

        report = {
            "country_count": len(zones),
            "small_rate_count": len(small_rates),
            "heavy_tier_count": len(heavy_tiers),
            "config": config,
        }

        return {
            "report": report,
            "zones": zones,
            "small_rates": small_rates,
            "heavy_tiers": heavy_tiers,
        }
    finally:
        wb.close()


if __name__ == "__main__":
    import json
    r = import_dhl()
    print(json.dumps(r["report"], indent=2))
