"""Phase F0: inspect freight workbook structure, headers, and key formula locations.

Reads D重量运费.xlsx and prints:
- Sheet list with dimensions
- Key header rows
- Formula cells (critical for understanding business logic)
- Country/zone distribution summary for each sheet

This script does NOT parse prices — that's the importer's job.
"""

from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = PROJECT_ROOT / "D重量运费.xlsx"

# ── helpers ──────────────────────────────────────

def _val(cell: Any) -> str:
    if cell is None:
        return "-"
    if isinstance(cell, float):
        if cell == int(cell):
            return str(int(cell))
        return f"{cell:.2f}"
    return str(cell).strip()

def _is_price(v: str) -> bool:
    try:
        f = float(v)
        return 0 < f < 100000
    except ValueError:
        return False

def _is_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")

# ── main inspection ──────────────────────────────

def inspect() -> None:
    print("=" * 72)
    print(f"Freight Workbook Inspection: {WORKBOOK.name}")
    print("=" * 72)

    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    try:
        sheet_titles = [ws.title for ws in wb.worksheets]
        print(f"\nSheets found: {len(wb.worksheets)}")
        for i, title in enumerate(sheet_titles, 1):
            ws = wb[title]
            print(f"  {i}. {title}  ({ws.max_row} rows × {ws.max_column} cols)")

        for ws in wb.worksheets:
            if not ws.title.lower().startswith(("区域运费", "区域", "2kg", "广诚", "体积公式", "物流询价")):
                continue

            print(f"\n{'─' * 72}")
            print(f"Sheet: {ws.title}")
            print(f"  Dimensions: {ws.max_row} rows × {ws.max_column} cols")
            print(f"  Min row: {ws.min_row}, Min col: {ws.min_column}")

            # ── Header rows (first 3) ──
            print(f"\n  Header rows:")
            for r in range(1, min(4, ws.max_row + 1)):
                row_vals = [_val(ws.cell(row=r, column=c).value)
                            for c in range(1, min(ws.max_column + 1, 80))]
                non_empty = [v for v in row_vals if v != "-"]
                print(f"    Row {r}: {', '.join(non_empty[:60])}")

            # ── Formula cells ──
            formulas = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
                for cell in row:
                    if _is_formula(cell):
                        formulas.append(f"{cell.coordinate}: {cell.value[:120]}")

            if formulas:
                print(f"\n  Formula cells ({len(formulas)} total, showing first 20):")
                for f in formulas[:20]:
                    print(f"    {f}")
                if len(formulas) > 20:
                    print(f"    ... and {len(formulas) - 20} more")

            # ── Country distribution (for rate sheets) ──
            if ws.title.lower() in {"区域运费dhl", "区域运费fedex"}:
                zones = defaultdict(int)
                countries_seen = 0
                for r in range(2, min(ws.max_row + 1, 500)):
                    country = _val(ws.cell(row=r, column=1).value)
                    if not country or country == "-":
                        break
                    zone = _val(ws.cell(row=r, column=2).value)
                    zones[zone] += 1
                    countries_seen += 1
                print(f"\n  Country count: {countries_seen}")
                print(f"  Zone distribution: {dict(sorted(zones.items()))}")

            # ── Weight headers (for rate sheets) ──
            if ws.title.lower().startswith("区域运费"):
                weights = []
                for c in range(3, min(ws.max_column + 1, 80)):
                    v = _val(ws.cell(row=1, column=c).value)
                    try:
                        w = float(v)
                        weights.append(w)
                    except ValueError:
                        pass
                print(f"  Weight columns ({len(weights)}): {weights}")

            # ── 2KG内文件 sheet ──
            if "2kg" in ws.title.lower() or "文件" in ws.title:
                print(f"\n  Document rate sheet inspection:")
                for r in range(1, min(ws.max_row + 1, 12)):
                    row_vals = [_val(ws.cell(row=r, column=c).value)
                                for c in range(1, min(ws.max_column + 1, 15))]
                    non_empty = [v for v in row_vals if v != "-"]
                    print(f"    Row {r}: {', '.join(non_empty)}")

            # ── 广诚 sheets: heavy cargo ──
            if "广诚" in ws.title.lower():
                # Scan for heavy cargo sections
                print(f"\n  Scanning for heavy cargo and rate tables...")
                heavy_header_found = False
                for r in range(1, ws.max_row + 1):
                    row_vals = [_val(ws.cell(row=r, column=c).value)
                                for c in range(1, min(ws.max_column + 1, 20))]
                    row_text = " ".join(row_vals).lower()
                    if any(kw in row_text for kw in ["重量段", "每kg", "重货", "单价", "包装重量", "heavy"]):
                        if not heavy_header_found:
                            print(f"\n  Heavy cargo section marker found at row {r}:")
                            heavy_header_found = True
                        # Print this row and next 3
                        for offset in range(4):
                            rr = r + offset
                            if rr > ws.max_row:
                                break
                            vals = [_val(ws.cell(row=rr, column=c).value)
                                    for c in range(1, min(ws.max_column + 1, 15))]
                            non_empty = [v for v in vals if v != "-"]
                            print(f"    Row {rr}: {', '.join(non_empty)}")

                # Find packaging weight formulas
                print(f"\n  Packaging weight formulas in this sheet:")
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 500)):
                    for cell in row:
                        if _is_formula(cell) and any(kw in str(cell.value).lower()
                                                      for kw in ["if", "packaging", "包装", "重量", "weight"]):
                            print(f"    {cell.coordinate}: {cell.value[:200]}")
                            # Print context rows
                            for ctx_r in range(max(1, cell.row - 2), min(ws.max_row + 1, cell.row + 3)):
                                ctx_vals = [_val(ws.cell(row=ctx_r, column=cc).value)
                                            for cc in range(1, min(ws.max_column + 1, 10))]
                                non_empty = [v for v in ctx_vals if v != "-"]
                                print(f"      Row {ctx_r}: {', '.join(non_empty)}")

    finally:
        wb.close()

    print(f"\n{'=' * 72}")
    print("Inspection complete.")
    print("=" * 72)


if __name__ == "__main__":
    inspect()
