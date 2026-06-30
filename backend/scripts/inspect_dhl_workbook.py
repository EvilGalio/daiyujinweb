"""Inspect A重量运费重制版.xlsx — DHL-only, single sheet."""

from pathlib import Path
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
WB = PROJECT_ROOT / "_private" / "source_data" / "freight" / "A重量运费重制版.xlsx"
if not WB.exists():
    WB = PROJECT_ROOT / "A重量运费重制版.xlsx"

def _num(v):
    if v is None: return None
    try: return round(float(v), 4)
    except: return None

def _text(v):
    return str(v).strip() if v is not None else ""

wb = load_workbook(WB, read_only=False, data_only=True)
try:
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"Sheet: {ws_name}  ({ws.max_row} rows x {ws.max_column} cols)")

        # Row 1: headers
        r1 = [_text(ws.cell(row=1, column=c).value) for c in range(1, min(ws.max_column+1, 70))]
        first_10 = [v for v in r1[:15] if v]
        print(f"  Row 1 first 15 non-empty: {first_10}")

        # A2:B5 (first countries + zones)
        print(f"  First countries:")
        for r in range(2, 6):
            a = _text(ws.cell(row=r, column=1).value)
            b = _text(ws.cell(row=r, column=2).value)
            if a:
                print(f"    Row {r}: {a}  zone={b}")

        # A232:A238 (last countries + heavy tier area)
        print(f"  Around row 232-238:")
        for r in range(232, 240):
            vals = [_text(ws.cell(row=r, column=c).value) for c in range(1, 11)]
            non_empty = [v for v in vals if v]
            print(f"    Row {r}: {non_empty[:8]}")

        # Heavy tier area: A235:J238
        print(f"  Heavy tier table A235:J238:")
        for r in range(235, 239):
            vals = [_text(ws.cell(row=r, column=c).value) for c in range(1, 11)]
            print(f"    Row {r}: {vals}")

        # Row 247-248: small cargo calc area
        print(f"  Small calc area row 247-248:")
        for r in range(247, 250):
            vals = [_text(ws.cell(row=r, column=c).value) for c in range(1, 9)]
            print(f"    Row {r}: {vals}")

        # Row 273-274: heavy calc area
        print(f"  Heavy calc area row 273-274:")
        for r in range(273, 276):
            vals = [_text(ws.cell(row=r, column=c).value) for c in range(1, 9)]
            print(f"    Row {r}: {vals}")

        # Country count (A2 to first empty or weight tier)
        count = 0
        for r in range(2, ws.max_row + 1):
            a = _text(ws.cell(row=r, column=1).value)
            if not a or 'kg' in a.lower():
                break
            # Check if it looks like a weight tier
            if any(ch in a for ch in ['-', '~']):
                try:
                    float(a.split('-')[0])
                    break
                except:
                    pass
            count += 1
        print(f"  Country count: {count}")

        # Last 3 non-empty rows in country list
        last_countries = []
        for r in range(2, count + 1):
            a = _text(ws.cell(row=r, column=1).value)
            if a:
                last_countries = (last_countries + [(r, a)])[-3:]
        print(f"  Last 3 countries: {last_countries}")

        # Golden case verification: Japan row, German row
        for target in ['日本', '德国']:
            for r in range(2, count + 2):
                if _text(ws.cell(row=r, column=1).value) == target:
                    zone = _text(ws.cell(row=r, column=2).value)
                    # Price at 6kg
                    for c in range(3, ws.max_column + 1):
                        w = _num(ws.cell(row=1, column=c).value)
                        if w and abs(w - 6.0) < 0.01:
                            price_6kg = _num(ws.cell(row=r, column=c).value)
                            print(f"  {target}: zone={zone}, 6kg price={price_6kg}")
                            break
                    break

finally:
    wb.close()
