"""Phase F0: extract golden cases directly from Excel for regression testing.

Reads D重量运费.xlsx with data_only=True to get computed values.
Produces 6 core golden cases and 1 advanced volumetric case.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = PROJECT_ROOT / "_private" / "source_data" / "freight" / "D重量运费.xlsx"
if not WORKBOOK.exists():
    WORKBOOK = PROJECT_ROOT / "D重量运费.xlsx"


def _num(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return round(float(v), 4)
    except (ValueError, TypeError):
        return None


def _text(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def find_country_in_sheet(ws, country_name: str, col_country=0, col_zone=1) -> tuple[int, str, int] | None:
    """Return (row_index, zone_code, data_start_col) or None."""
    for row_idx in range(2, ws.max_row + 1):
        c = _text(ws.cell(row=row_idx, column=col_country + 1).value)
        if country_name.lower() in c.lower():
            z = _text(ws.cell(row=row_idx, column=col_zone + 1).value)
            return row_idx, z, col_country + 2
    return None


def get_weight_headers(ws, start_col: int) -> list[tuple[int, float]]:
    """Return [(col_index, weight_kg), ...] from header row."""
    result = []
    for c in range(start_col, ws.max_column + 1):
        v = _num(ws.cell(row=1, column=c).value)
        if v is not None and v > 0:
            result.append((c, v))
    return result


def find_price_at_weight(ws, row: int, weight_headers: list[tuple[int, float]], target_kg: float) -> tuple[float, float] | None:
    """Return (matched_weight, price) for the weight tier >= target_kg."""
    for col, w in weight_headers:
        if w >= target_kg:
            p = _num(ws.cell(row=row, column=col).value)
            if p is not None:
                return w, p
    return None


def extract_golden_cases() -> dict[str, Any]:
    wb = load_workbook(WORKBOOK, read_only=False, data_only=True)
    try:
        cases = {}

        # ── DHL matrix zone lookup ──
        ws_dhl = wb["区域运费DHL"]
        dhl_wh = get_weight_headers(ws_dhl, start_col=3)

        # ── FedEx matrix zone lookup ──
        ws_fex = wb["区域运费FedEX"]
        fex_wh = get_weight_headers(ws_fex, start_col=3)

        # ── 2KG内文件 lookup ──
        ws_doc = wb["2KG内文件"]
        doc_weights = {0.5: 4, 1.0: 5, 1.5: 6, 2.0: 7}
        doc_zones = {f"{i}区": i for i in range(1, 10)}

        # ── 广诚DHL heavy cargo ──
        ws_gc_dhl = wb["广诚DHL运费"]
        # =RMB*1.2/6.5*1.2 for USD columns (O-W)
        # =RMB*1.2/7*1.2 for EUR columns (AB-AJ)
        # Heavy at rows 66-68

        # ── 广诚FedEx heavy cargo ──
        ws_gc_fex = wb["广诚FedEx运费"]
        # Heavy at rows 46+

        # ────────────────────────────────────────────
        # Golden Case 1: DHL Germany 5kg small parcel
        # ────────────────────────────────────────────
        de_dhl = find_country_in_sheet(ws_dhl, "德国") or find_country_in_sheet(ws_dhl, "Germany")
        if de_dhl:
            row, zone, dc = de_dhl
            w_match, price = find_price_at_weight(ws_dhl, row, dhl_wh, 5.0) or (None, None)
            cases["gc1_dhl_germany_5kg"] = {
                "description": "DHL small parcel: Germany, 5.0kg, CNY, default USD display",
                "sheet": "区域运费DHL",
                "row": row,
                "zone": zone,
                "actual_weight_kg": 5.0,
                "billable_weight_kg": w_match,
                "base_price_cny": price,
                "pricing_mode": "small_matrix",
                "source": f"区域运费DHL row {row} 5kg col",
            }

        # ────────────────────────────────────────────
        # Golden Case 2: FedEx Germany 5kg small parcel
        # ────────────────────────────────────────────
        de_fex = find_country_in_sheet(ws_fex, "德国") or find_country_in_sheet(ws_fex, "Germany")
        if de_fex:
            row, zone, dc = de_fex
            w_match, price = find_price_at_weight(ws_fex, row, fex_wh, 5.0) or (None, None)
            cases["gc2_fedex_germany_5kg"] = {
                "description": "FedEx small parcel: Germany, 5.0kg, CNY, default USD display",
                "sheet": "区域运费FedEX",
                "row": row,
                "zone": zone,
                "actual_weight_kg": 5.0,
                "billable_weight_kg": w_match,
                "base_price_cny": price,
                "pricing_mode": "small_matrix",
                "source": f"区域运费FedEX row {row} 5kg col",
            }

        # ────────────────────────────────────────────
        # Golden Case 3: DHL Document 1.5kg
        # ────────────────────────────────────────────
        # Find country in DHL zone 5 (doesn't matter which — just needs a zone)
        # For simplicity, use zone 1 example
        doc_prices = {}
        for wt, row_num in doc_weights.items():
            for zone_str, zone_num in doc_zones.items():
                col = zone_num + 1  # +1 because col A is "KG"
                p = _num(ws_doc.cell(row=row_num, column=col).value)
                if p is not None:
                    doc_prices[f"zone_{zone_num}_{wt}kg"] = p

        cases["gc3_dhl_document_1_5kg"] = {
            "description": "DHL document rate: any country, 1.5kg, CNY",
            "sheet": "2KG内文件",
            "document_prices_sample": doc_prices,
            "weight": 1.5,
            "expected_zone_price_entry": doc_prices.get("zone_5_1.5kg", None),
            "pricing_mode": "document",
        }

        # ────────────────────────────────────────────
        # Golden Case 4: DHL 36kg heavy cargo
        # ────────────────────────────────────────────
        # Find heavy tier in 广诚DHL运费 rows 66-68
        dhl_heavy = {}
        heavy_headers_dhl = [_text(ws_gc_dhl.cell(row=65, column=c).value)
                             for c in range(1, 13)]
        for r in [66, 67, 68]:
            lo = _num(ws_gc_dhl.cell(row=r, column=1).value)
            hi = _num(ws_gc_dhl.cell(row=r, column=2).value)
            if lo is None:
                continue
            zone_prices = {}
            for zi in range(1, 10):
                p = _num(ws_gc_dhl.cell(row=r, column=2 + zi).value)
                if p is not None:
                    zone_prices[f"zone_{zi}"] = p
            dhl_heavy[f"{lo}-{hi}"] = zone_prices

        # 36kg falls in 30.1-70 tier
        unit_price_zone5 = dhl_heavy.get("30.1-70.0", {}).get("zone_5")
        packaging_rule = 36 + 7.5  # = 43.5

        cases["gc4_dhl_heavy_36kg"] = {
            "description": "DHL heavy cargo: 36kg, zone 5, packaging +7.5kg",
            "sheet": "广诚DHL运费",
            "actual_weight_kg": 36.0,
            "packaging_rule": "+7.5kg",
            "adjusted_weight_kg": packaging_rule,
            "heavy_tier": "30.1-70",
            "unit_price_per_kg_cny": unit_price_zone5,
            "expected_base_cny": round(unit_price_zone5 * packaging_rule, 2) if unit_price_zone5 else None,
            "pricing_mode": "heavy_per_kg",
            "all_heavy_tiers": {k: v for k, v in dhl_heavy.items() if "zone_5" in v},
        }

        # ────────────────────────────────────────────
        # Golden Case 5: FedEx 25kg heavy cargo
        # ────────────────────────────────────────────
        fex_heavy = {}
        # Heavy section starts at row 46
        for r in range(46, ws_gc_fex.max_row + 1):
            lo_text = _text(ws_gc_fex.cell(row=r, column=1).value)
            if not lo_text or "重货" in lo_text:
                continue
            parts = [x.strip() for x in lo_text.replace("-", " ").replace("–", " ").split()]
            try:
                lo = float(parts[0])
                hi = float(parts[1]) if len(parts) > 1 else None
            except (ValueError, IndexError):
                continue
            if lo is None:
                continue
            zone_prices = {}
            # Columns B onwards are zone prices
            for ci in range(2, min(ws_gc_fex.max_column + 1, 15)):
                p = _num(ws_gc_fex.cell(row=r, column=ci).value)
                if p is not None:
                    # Map to zone letter from header row 3
                    zone_letter = _text(ws_gc_fex.cell(row=3, column=ci).value)
                    zone_prices[zone_letter or f"col_{ci}"] = p
            fex_heavy[f"{lo}-{hi}"] = zone_prices

        # 25kg falls in 21-44 tier. Look for first tier price
        tier_21_44 = fex_heavy.get("21.0-44.0", {})
        sample_price = next(iter(tier_21_44.values())) if tier_21_44 else None
        sample_zone = next(iter(tier_21_44.keys())) if tier_21_44 else None
        packaging_fex = 25 + 4  # = 29

        cases["gc5_fedex_heavy_25kg"] = {
            "description": "FedEx heavy cargo: 25kg, packaging +4kg",
            "sheet": "广诚FedEx运费",
            "actual_weight_kg": 25.0,
            "packaging_rule": "+4kg",
            "adjusted_weight_kg": packaging_fex,
            "heavy_tier": "21-44",
            "sample_zone": sample_zone,
            "sample_unit_price_cny": sample_price,
            "expected_base_cny": round(sample_price * packaging_fex, 2) if sample_price else None,
            "pricing_mode": "heavy_per_kg",
            "all_heavy_tiers": fex_heavy,
        }

        # ────────────────────────────────────────────
        # Golden Case 6: Billing weight rounding rules
        # ────────────────────────────────────────────
        # From 广诚DHL运费 row 1 text
        cases["gc6_billing_weight_rules"] = {
            "description": "DHL billing weight rounding rules (from 广诚DHL运费 row 1)",
            "rules": {
                "min": 0,
                "max_5kg": "round up by 1-1.5kg",
                "5_10kg": "round up by 2kg",
                "10_20kg": "round up by 3kg",
                "20_30kg": "use 10-20kg pricing (cap billing weight at 20kg, use 20kg tier price)",
                "heavy_30_70kg": "round up by 5-10kg before packaging",
                "heavy_70_300kg": "round up by 10-15kg before packaging",
                "heavy_over_300kg": "weight * 1.0822 (from formula in Excel)",
            },
            "test_cases_billing_weight": {
                "4.2kg": "round up to 5kg or 5.5kg",
                "7.0kg": "round up to 9kg",
                "12.0kg": "round up to 15kg",
                "25.0kg": "use 20kg tier pricing (cap at 20kg billing weight)",
            },
        }

        # ────────────────────────────────────────────
        # Advanced: Volumetric weight case
        # ────────────────────────────────────────────
        cases["gc7_volumetric"] = {
            "description": "Advanced: volumetric weight exceeds actual weight",
            "dimensions": {"length_cm": 50, "width_cm": 40, "height_cm": 30, "boxes": 1},
            "divisor": 5000,
            "volumetric_weight_kg": round(50 * 40 * 30 / 5000, 2),  # = 12.0
            "actual_weight_kg": 5.0,
            "billable_base_weight_kg": 12.0,
            "note": "volumetric weight > actual, use volumetric as billable",
        }

        return cases
    finally:
        wb.close()


if __name__ == "__main__":
    cases = extract_golden_cases()
    print(json.dumps(cases, indent=2, ensure_ascii=False))
